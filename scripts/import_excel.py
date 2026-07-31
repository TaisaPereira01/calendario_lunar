"""
Protocolos Lunares

Importa diretamente o arquivo Excel para o SQLite.

Pipeline

Excel
    ↓
SQLite
    ↓
Streamlit
"""

from pathlib import Path
import sqlite3
import re

from openpyxl import load_workbook


# =============================================================================
# PATHS
# =============================================================================

ROOT = Path(__file__).resolve().parents[1]

DATA_DIR = ROOT / "data"

DATABASE_DIR = ROOT / "database"

DATABASE = DATABASE_DIR / "protocolos.db"

EXCEL = DATA_DIR / "Calendario_Lunar_Integrativo.xlsx"

MOON_CALENDAR = DATA_DIR / "moon_calendar.json"


# =============================================================================
# WORKBOOK
# =============================================================================

SHEETS = [

    "Lua Nova",

    "Lua Crescente",

    "Lua Cheia",

    "Lua Minguante",

]


# =============================================================================
# DIAS
# =============================================================================

WEEKDAYS = {

    3: "segunda",

    4: "terça",

    5: "quarta",

    6: "quinta",

    7: "sexta",

    8: "sábado",

    9: "domingo",

}


# =============================================================================
# LINHAS
# =============================================================================

PERIODS = {

    7: ("Rotina Matinal", "ROUTINE"),

    8: ("Café da Manhã", "FOOD"),

    9: ("Suplementos Manhã", "SUPPLEMENT"),

    10: ("Almoço", "FOOD"),

    11: ("Suplementos Tarde", "SUPPLEMENT"),

    12: ("Lanche", "FOOD"),

    13: ("Jantar", "FOOD"),

    14: ("Antes de Dormir", "ROUTINE"),

    15: ("Exercício", "EXERCISE"),

    16: ("Terapias", "THERAPY"),

}


# =============================================================================
# CACHE
# =============================================================================

CACHE = {

    "phase": {},

    "weekday": {},

    "period": {},

    "item_type": {},

    "item": {},

}


# =============================================================================
# DATABASE
# =============================================================================

connection = sqlite3.connect(DATABASE)

connection.row_factory = sqlite3.Row

cursor = connection.cursor()


# =============================================================================
# SQL
# =============================================================================

def execute(sql, params=()):

    cursor.execute(sql, params)

    return cursor


# =============================================================================
# HELPERS
# =============================================================================

def clean(value):

    if value is None:

        return ""

    text = str(value)

    text = text.replace("\r", "\n")

    text = text.replace("•", "\n")

    text = re.sub(r"\n+", "\n", text)

    return text.strip()


def split_lines(text):

    text = clean(text)

    if not text:

        return []

    return [

        line.strip()

        for line in text.split("\n")

        if line.strip()

    ]


# =============================================================================
# CACHE
# =============================================================================

def load_cache():

    CACHE["phase"].clear()

    CACHE["weekday"].clear()

    CACHE["period"].clear()

    CACHE["item_type"].clear()

    CACHE["item"].clear()

    #
    # Phase
    #

    for row in execute(

        "SELECT id,name FROM phase"

    ).fetchall():

        CACHE["phase"][row["name"]] = row["id"]

    #
    # Weekday
    #

    for row in execute(

        "SELECT id,name FROM weekday"

    ).fetchall():

        CACHE["weekday"][row["name"]] = row["id"]

    #
    # Period
    #

    for row in execute(

        "SELECT id,name FROM period"

    ).fetchall():

        CACHE["period"][row["name"]] = row["id"]

    #
    # Item Type
    #

    for row in execute(

        "SELECT id,name FROM item_type"

    ).fetchall():

        CACHE["item_type"][row["name"]] = row["id"]

    #
    # Item
    #

    rows = execute(

        """

        SELECT

            i.id,

            t.name AS type,

            i.name

        FROM item i

        JOIN item_type t

        ON t.id=i.item_type_id

        """

    ).fetchall()

    for row in rows:

        CACHE["item"][

            (

                row["type"],

                row["name"]

            )

        ] = row["id"]


# =============================================================================
# ITEM
# =============================================================================

def get_or_create_item(

    item_type,

    name,

    description=""

):

    key = (

        item_type,

        name,

    )

    if key in CACHE["item"]:

        return CACHE["item"][key]

    execute(

        """

        INSERT INTO item

        (

            item_type_id,

            name,

            description

        )

        VALUES

        (?,?,?)

        """,

        (

            CACHE["item_type"][item_type],

            name,

            description,

        ),

    )

    item_id = cursor.lastrowid

    CACHE["item"][key] = item_id

    return item_id

# =============================================================================
# CELL PARSER
# =============================================================================

# Uma linha inteira entre parênteses é um qualificador do item anterior,
# não um item próprio — ex.: "(Sempre com alguma gordura)" pertence à
# Vitamina D3 que vem antes. Vira "notes" do item anterior. Ver BUG-001.

PAREN_NOTE = re.compile(r"^\((.+)\)$")


def parse_cell(value, item_type):
    """
    Converte uma célula em uma lista de itens.

    Cada linha não vazia da célula gera um item.

    Exceção: uma linha totalmente entre parênteses não é um item, e sim
    uma nota (qualificador) do item imediatamente anterior. Ver BUG-001.

    Exemplo:

        Água morna
        Limão
        Respiração

    →

    [
        {...},
        {...},
        {...}
    ]
    """

    items = []

    for line in split_lines(value):

        note = PAREN_NOTE.match(line)

        if note and items:

            inner = note.group(1).strip()

            previous = items[-1]

            previous["notes"] = (

                f'{previous["notes"]} {inner}'.strip()

                if previous["notes"]

                else inner

            )

            continue

        items.append({

            "type": item_type,

            "name": line,

            "value": "",

            "description": "",

            "notes": ""

        })

    return items


# =============================================================================
# SHEET
# =============================================================================

def parse_sheet(sheet_name):

    print(f"Importando aba: {sheet_name}")

    ws = workbook[sheet_name]

    protocol = {

        "phase": sheet_name,

        "days": {}

    }

    #
    # Inicializa estrutura
    #

    for weekday in WEEKDAYS.values():

        protocol["days"][weekday] = {}

    #
    # Percorre linhas (períodos)
    #

    for row, (period_name, item_type) in PERIODS.items():

        #
        # Inicializa período
        #

        for weekday in WEEKDAYS.values():

            protocol["days"][weekday][period_name] = []

        #
        # Percorre dias
        #

        for column, weekday in WEEKDAYS.items():

            value = ws.cell(

                row=row,

                column=column

            ).value

            if value is None:

                continue

            items = parse_cell(

                value,

                item_type

            )

            protocol["days"][weekday][period_name].extend(

                items

            )

    return protocol


# =============================================================================
# MOON CALENDAR
# =============================================================================

def import_moon_calendar():

    import json

    execute(

        "DELETE FROM moon_calendar"

    )

    with open(

        MOON_CALENDAR,

        encoding="utf-8"

    ) as fp:

        calendar = json.load(fp)

    for row in calendar:

        execute(

            """

            INSERT INTO moon_calendar

            (

                date,

                phase_id

            )

            VALUES

            (?,?)

            """,

            (

                row["date"],

                CACHE["phase"][

                    row["phase"]

                ]

            )

        )

    connection.commit()


# =============================================================================
# WORKBOOK
# =============================================================================

print("Abrindo Excel...")

workbook = load_workbook(

    EXCEL,

    data_only=True

)

# =============================================================================
# GENERATOR
# =============================================================================

def iter_sheet(sheet_name):

    print(f"Importando {sheet_name}")

    ws = workbook[sheet_name]

    for row, (period_name, item_type) in PERIODS.items():

        for column, weekday in WEEKDAYS.items():

            value = ws.cell(

                row=row,

                column=column

            ).value

            if value is None:

                continue

            items = parse_cell(

                value,

                item_type

            )

            for item in items:

                yield (

                    sheet_name,

                    weekday,

                    period_name,

                    item,

                )


# =============================================================================
# PROTOCOL
# =============================================================================

def insert_protocol_item(

    phase,

    weekday,

    period,

    item,

    order,

):

    execute(

        """

        INSERT INTO protocol_item

        (

            phase_id,

            weekday_id,

            period_id,

            item_id,

            display_order,

            value,

            notes

        )

        VALUES

        (?,?,?,?,?,?,?)

        """,

        (

            CACHE["phase"][phase],

            CACHE["weekday"][weekday],

            CACHE["period"][period],

            get_or_create_item(

                item["type"],

                item["name"]

            ),

            order,

            item.get(

                "value",

                ""

            ),

            item.get(

                "notes",

                ""

            ),

        )

    )


# =============================================================================
# IMPORT
# =============================================================================

def import_protocols():

    execute(

        "DELETE FROM protocol_item"

    )

    connection.commit()

    total = 0

    for sheet in SHEETS:

        order = 1

        for (

            phase,

            weekday,

            period,

            item,

        ) in iter_sheet(sheet):

            insert_protocol_item(

                phase,

                weekday,

                period,

                item,

                order,

            )

            total += 1

            order += 1

    connection.commit()

    return total

# =============================================================================
# MAIN
# =============================================================================

def main():

    print("=" * 60)
    print("Protocolos Lunares")
    print("=" * 60)

    #
    # Atualiza caches
    #

    load_cache()

    try:

        connection.execute("BEGIN")

        #
        # Calendário Lunar
        #

        import_moon_calendar()

        #
        # Protocolos
        #

        execute("DELETE FROM protocol_item")

        orders = {}

        total = 0

        for sheet in SHEETS:

            for (

                phase,

                weekday,

                period,

                item,

            ) in iter_sheet(sheet):

                key = (

                    phase,

                    weekday,

                    period,

                )

                order = orders.get(

                    key,

                    0,

                ) + 1

                orders[key] = order

                insert_protocol_item(

                    phase,

                    weekday,

                    period,

                    item,

                    order,

                )

                total += 1

        connection.commit()

    except Exception:

        connection.rollback()

        raise

    print()

    print("=" * 60)

    print("Importação concluída")

    print("=" * 60)

    print(f"Protocolos importados : {total}")

    print("=" * 60)

    connection.close()


# =============================================================================

if __name__ == "__main__":

    main()

